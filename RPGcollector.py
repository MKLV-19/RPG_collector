# -*- coding: utf-8 -*-
"""
Created on Fri Nov  1 21:12:40 2024

@author: 切糕

intro: 将当前目录中的统计表和说明文档组成排行榜图像的程序
"""
# 包管理
# 加载包
import os
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook
import textwrap
import datetime
import glob
import configparser

# 文件名称预设
#工作路径
cwpath=os.path.dirname(os.path.abspath(__file__)) #读取文件所在路径
#读取
config_path=os.path.join(cwpath,'config.ini')
data_path=glob.glob(os.path.join(cwpath, '*统计*.xlsx'))[0] #读取表格的名称（允许通配）
text_file_path = glob.glob(os.path.join(cwpath,'*说明*.txt'))[0]   #读取文字说明的路径（允许通配）
font_path = os.path.join(cwpath,'萝莉体.ttf')  # 指定字体文件路径
#config读取
config = configparser.ConfigParser()
config.read(config_path)
#输出
output_folder = os.path.join(cwpath, 'output') #输出文件的路径
os.makedirs(output_folder, exist_ok=True)   #如果没有文件夹，就创建这个文件夹
pic_name=config['DEFAULT']['output_img_name'] +'.'+config['DEFAULT']['output_img_ext']
pic_path=os.path.join(output_folder, pic_name) #输出图像的路径和名称


# 图像预设（图片以及图中所有元素的大小均取决于icon的尺寸）
##核心参数
icon_size = [int(config['DEFAULT']['icon_size_width']),int(config['DEFAULT']['icon_size_height']) ] #单个图标的宽和高
icon_num_perline = int(config['DEFAULT']['icon_num_perline'])  #每行出现的图像数量
title=config['DEFAULT']['title']     #图片标题
date='更新日期: '+ config['DEFAULT']['update_date'] 
# icon_size = [100,100]   #单个图标的宽和高
# icon_num_perline = 8    #每行出现的图像数量
# title='RPG战力榜（个人向）'     #图片标题
# date='更新日期: '+ str(datetime.date.today())   #更新日期
##衍生参数
#图片间隔
icon_interval_width=round(icon_size[0]/10)  #图片之间的空隙宽度
icon_interval_height=round(icon_size[1]/10) #图片之间的空隙高度
#盒子
box_width=icon_num_perline*(icon_size[0]+icon_interval_width)   #一组图片的宽度
box_information_width=box_width+icon_size[0]
box_interval_height=round(icon_size[1]/5)   #两组图片的空隙高度
#画布
canvas_title_height=round(icon_size[1]/2)   #标题大小
canvas_date_height=round(icon_size[1]/4)    #日期大小
canvas_top_height=round(icon_size[1]/5)    #页面顶部空白大小
canvas_bottom_height=round(icon_size[1]/5)    #页面底部空白大小
canvas_left_width=icon_size[0]    #页面左侧空白大小
canvas_right_width=round(icon_size[0]/2)  #页面右侧空白大小
#字体
font_size_title=canvas_title_height  #标题字体大小
font_size_date=canvas_date_height    #日期字体大小
font_size_rank=round(canvas_left_width/3)  #排名字体大小
font_size_information=round(icon_size[0]/4) #说明字体大小
font_interval_height=round(font_size_information/10)    #字体行间距
#坐标
position_rank_x=round(canvas_left_width/3)  #排名横坐标
position_information_x=round(canvas_left_width/2)    #说明横坐标
#高度修正
delta_height=round(icon_size[1]/10)


# 函数
#读取excel中的图片信息
def get_dict_from_excel(excel_path=data_path):
    '''
    返回一个由rank和reflist构成的字典
    从指定路径中获取xlsx文件中的数据，包含图片，行索引，对应等级
    按图片对应等级将数据进行划分并根据索引进行分组排序，转化为以等级为键、图片列表为值的字典后返回
    '''
    #文件读取
    workbook=load_workbook(filename=excel_path,data_only=True)  #使用workbook读取xlsx文件的信息
    worksheet=workbook['Sheet1']
    
    #读取评分所在行列
    found_cells=[]
    founded_flag=False
    for row in worksheet.iter_rows():
        if founded_flag:break
        for cell in row:
            if '评分' in str(cell.value):
                found_cells.append((cell.row, cell.column))
                founded_flag=True
                break
    rankcol=found_cells[0][1]
    
    #获取工作表中的所有图像
    images=worksheet._images  
    
    #获取图片相关信息
    index=[]
    img_ref=[]
    img_rank=[]
    curr_row=0
    for image in images:
        curr_row=image.anchor._from.row+1
        index=index+[curr_row]  #每张图片对应excel表中的行索引
        img_ref=img_ref+[image.ref] #每张图片对应的BytesIO数据，用于绘制图像
        img_rank=img_rank+[worksheet.cell(curr_row,rankcol).value]  #每张图片对应的等级
    
    #将得到的信息列表合并
    assert len(index)==len(img_ref)==len(img_rank), "所有列表的长度必须相同"
    df=pd.DataFrame(list(zip(index,img_ref,img_rank)),columns=['index', 'ref', 'rank']) #合并数据为一个dataframe
    
    #数据的分组
    df=df.sort_values(by='rank') #按照等级将数据排序
    grouped_df=df.groupby(by='rank') #按照等级对数据进行分组
    
    #将分组后的数据按照索引进行组内排序，并以字典形式输出
    sorted_groups = {rank: group.sort_values(by='index') for rank, group in grouped_df}
    return sorted_groups  

def get_dict_from_txt(text_file_path=text_file_path,
                      font_size_information=font_size_information,
                      box_information_width=box_information_width,
                      font_interval_height=font_interval_height):
    '''
    从指定路径读取txt文件，并根据盒子的宽度以及字体大小返回：单行文本高度，文本框高度，文本列表
    '''
    #读取txt文件
    with open(text_file_path, 'r', encoding='utf-8') as file:
        text = file.read()

    #将文本写入文本框
    lines = text.split('\n') #使用'\n'切割文本，使得文本在换行处被分开
    wrapped_lines = []  #将字符串写入可换行的文本框中
    for line in lines:
        wrapped_lines.extend(textwrap.wrap(line, width=box_information_width // font_size_information))
    line_height = font_size_information + font_interval_height #单行高度+行间距
    total_height = (len(wrapped_lines)+1) * line_height #总高度
    return {'line_height':line_height,
            'total_height':total_height,
            'texts':wrapped_lines}

def get_img_from_dict_and_text(img_dict,text_dict,title=title,date=date,
                          icon_size=icon_size,
                          icon_num_perline=icon_num_perline,
                          icon_interval_height=icon_interval_height,
                          icon_interval_width=icon_interval_width,
                          box_width=box_width,
                          box_information_width=box_information_width,
                          box_interval_height=box_interval_height,
                          canvas_title_height=canvas_title_height,
                          canvas_date_height=canvas_date_height,
                          canvas_top_height=canvas_top_height,
                          canvas_bottom_height=canvas_bottom_height,
                          canvas_left_width=canvas_left_width,
                          canvas_right_width=canvas_right_width,
                          font_size_title=font_size_title,
                          font_size_date=font_size_date,
                          font_size_rank=font_size_rank,
                          font_size_information=font_size_information,
                          font_interval_height=font_interval_height,
                          position_rank_x=position_rank_x,
                          position_information_x=position_information_x,
                          delta_height=delta_height,
                          font_path=font_path):
    '''
    将完成分类的图片字典和说明字典根据图像的相关参数生成图像盒子，返回一个合成后的图像
    '''
    #统计画布的高度
    canvas_height=0
    #页眉高度
    canvas_height=canvas_height+canvas_top_height+canvas_bottom_height
    #标题和日期高度
    canvas_height=canvas_height+canvas_title_height+canvas_date_height
    #遍历img_dict,得到图片表格高度
    for rank,group in img_dict.items():
        canvas_height=canvas_height+((len(group)-1)//icon_num_perline + 1)*icon_size[1]+((len(group)-1)//icon_num_perline)*icon_interval_height+box_interval_height
    # (这里采用(len(group)-1)计算是为了防止组中元素正好等于行图标个数整数倍时多计算一行)
    #说明文字总高度
    canvas_height=canvas_height+text_dict['total_height']
    #标题和日期的高度修正
    canvas_height=canvas_height+delta_height*2
    
    #统计画布的宽度
    canvas_width=0
    #页边缘宽度
    canvas_width=canvas_width+canvas_left_width+canvas_right_width
    #中心盒子宽度
    canvas_width=canvas_width+box_width
        
    #创建空白画布
    canvas = Image.new('RGB', (canvas_width, canvas_height), color='white')
    
    #根据坐标将各个元素写入画布
    current_pos=(0,0)   #当前坐标(x,y)
    
    #定义函数用以计算文本框宽度和高度
    def get_textbox_size(string,font):
        '''
        输入字符串，以及ImageFont.truetype，返回一段文本宽和高的二元组(width,height)
        '''
        draw=ImageDraw.Draw(Image.new('RGB', (1,1))) # 使用 ImageDraw 来计算文本的尺寸
        left, top, right, bottom = draw.textbbox((0, 0), string, font)
        return (right - left,bottom - top + delta_height) #加入修正防止高度不够
    
    #定义函数用以获取文本框图像
    def get_textimg(text,text_width,text_height,font):
        '''
        输入文本，文本框的宽和高，以及ImageFont.truetype，返回含有对应文本的图像
        '''
        image_text=Image.new('RGB',(text_width,text_height),color='white')
        draw=ImageDraw.Draw(image_text)
        draw.text((0, 0), text, font=font, fill=(0,0,0))
        return image_text
    
    #写入标题
    #加载字体
    font_title = ImageFont.truetype(font_path, font_size_title) 
    #计算标题尺寸
    title_width,title_height=get_textbox_size(title, font_title)
    #计算标题左上角坐标
    current_pos=(round((canvas_width-title_width)/2),canvas_top_height)
    #画出对应文本框
    image_title=get_textimg(title, title_width, title_height, font_title)
    #粘贴至画布指定位置
    canvas.paste(image_title, current_pos)
    
    #写入日期
    #加载字体
    font_date=ImageFont.truetype(font_path, font_size_date)    
    #计算日期尺寸
    date_width, date_height=get_textbox_size(date, font_date)
    #计算日期左上角坐标
    current_pos=(round(canvas_width-canvas_right_width/2-date_width),current_pos[1]+title_height)
    #画出对应文本框
    image_date=get_textimg(date, date_width, date_height, font_date)
    #粘贴至画布指定位置
    canvas.paste(image_date, current_pos)
    
    #写入图片列表
    #加载字体
    font_rank=ImageFont.truetype(font_path,font_size_rank)
    #计算当前坐标
    current_pos=(position_rank_x,current_pos[1]+date_height)
    #当前组所处高度
    current_group_height=current_pos[1]
    #遍历img_dict，分组将等级和图片写入画布
    for rank, group in img_dict.items():
        #将等级写入指定位置
        rank_width, rank_height=get_textbox_size(rank, font_rank)
        image_rank=get_textimg(rank, rank_width, rank_height, font_rank)
        canvas.paste(image_rank,current_pos)
        #遍历group中每一个图片，并将图片顺次写入指定位置
        current_num=0 #记录当前组图标的个数，如果在换行时正好是本组最后一个图标，则跳过坐标重定位
        icon_num=0  #当前行的图片数量记录，当达到最大数量时，需要刷新为0
        for index, row in group.iterrows():
            current_pos=(canvas_left_width+icon_num*(icon_size[0]+icon_interval_width),current_pos[1])
            #写入图片
            image_icon = Image.open(row['ref']).convert("RGB")  #根据ref得到图片
            #根据icon_size等比缩放图片
            target_width=icon_size[0]
            target_height=round(image_icon.size[1]*target_width/(image_icon.size[0]+0.01)) 
            image_icon = image_icon.resize((target_width, target_height)) 
            #根据icon_size在中心处剪切图片
            min_height=min(target_height,icon_size[1])
            upper=round((target_height-min_height)/2)
            lower=round((target_height+min_height)/2)
            image_icon = image_icon.crop((0, upper, target_width, lower))
            #粘贴图片
            canvas.paste(image_icon,current_pos)
            #图片数量更新
            icon_num=icon_num+1
            current_num=current_num+1
            if icon_num >= icon_num_perline:
                icon_num = 0    #换行
                if current_num >= len(group): 
                    break #如果在换行时正好是本组最后一个图标，则跳过坐标重定位
                current_pos=(canvas_left_width,current_pos[1]+icon_size[1]+icon_interval_height)
        #计算当前组所处高度
        current_group_height=current_group_height+((len(group)-1)//icon_num_perline + 1)*icon_size[1]+((len(group)-1)//icon_num_perline)*icon_interval_height+box_interval_height
        #等级坐标重定位
        current_pos=(position_rank_x,current_group_height)
        
    #写入说明文字
    #坐标重定位
    current_pos=(position_information_x,current_pos[1])
    #获得切割完成的文本及其相关高度
    wrapped_lines = text_dict['texts']
    total_height = text_dict['total_height']
    line_height = text_dict['line_height']
    #加载字体
    font_information = ImageFont.truetype(font_path, font_size_information) 
    #将文本逐行写入图片
    image_information = Image.new('RGB', (box_information_width, total_height),color='white')
    draw = ImageDraw.Draw(image_information)
    line_height_sum = 0 #行高记录
    for line in wrapped_lines:
        draw.text((0, line_height_sum), line, font=font_information, fill=(255,0,0))
        line_height_sum = line_height_sum + line_height
    #把文字粘贴到画布的指定位置
    canvas.paste(image_information, current_pos)
    
    #返回最终图片
    return canvas


# 运行
img_dict=get_dict_from_excel()
text_dict=get_dict_from_txt()
final_img=get_img_from_dict_and_text(img_dict, text_dict)
final_img.save(pic_path);
print("图像已保存至当前目录下的output文件夹")
